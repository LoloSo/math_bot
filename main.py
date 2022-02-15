import random
from settings import TOKEN
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler, CallbackContext, CallbackQueryHandler
from openpyxl import load_workbook
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from sys import exc_info
from traceback import extract_tb
import theory_text
from operator import itemgetter

wb = load_workbook('database.xlsx')
vendor = wb['users']

keyboard = [
    ["Теория"],
    ["Практика"],
    ['Узнать статистику', 'Сообщить о проблеме']
]

def main():

    updater = Updater(token=TOKEN, use_context=True)
    dispatcher = updater.dispatcher

    handler = MessageHandler(Filters.command, do_command)
    keyboard_handler = MessageHandler(Filters.text, keyboard_value)


    dispatcher.add_handler(handler)
    dispatcher.add_handler(keyboard_handler)


    updater.start_polling()
    updater.idle()


def do_start(update, context):

    keyboard = [
        ["Теория"],
        ["Практика"],
        ['Узнать статистику', 'Сообщить о проблеме']
    ]

    update.message.reply_text(text='Добро пожаловать, выберите, что бы вы хотели сделать.',
                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))


def keyboard_value(update: Update, context):
    text = update.message.text

    try:

        if 'command' in context.user_data:
            if update.message.text != 'Сообщить о проблеме' and context.user_data['command'] == '/addven':
                do_words(update, context)

        if text == 'Назад':
            context.user_data['sort'] = None
            context.user_data['practice'] = None
            do_start(update=update, context=context)
            return

        if text == 'Топ-5 пользователей':
            do_show_top_statistic(update=update, context=context)

        if text == 'Узнать статистику':
            user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
            do_show_statistic(name=user_name, update=update, context=context)
            return

        if text == 'Сообщить о проблеме':
            do_command(update=update, context=context)
            return

        if text == "Теория" or text == 'Назад' and context.user_data['back'] == 'назад теория':
            context.user_data['sort'] = 'Теория'
            keyboard = [
                ["Умножения двоичных чисел"],
                ["Возведение в квадрат"],
                ["Квадратные уравнения"],
                ["Назад"]
            ]
            update.message.reply_text(text='Выберите тему',reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))

        elif text == "Практика":
            context.user_data['sort'] = 'Практика'
            keyboard = [
                ["Умножения двоичных чисел", "Возведение в квадрат"],
                ["Квадратные уравнения", "Поиск процента"],
                ["Назад"]
            ]
            update.message.reply_text(text='Выберите тему',reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))

        check_practice = 'practice' in context.user_data

        if 'sort' in context.user_data:

            if update.message.text != 'Теория' and context.user_data['sort'] == 'Теория':

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Умножения двоичных чисел[3]':

                    context.user_data[context.user_data['sort']] = 'Умножения двоичных чисел[4]'
                    context.user_data['back'] = 'назад теория'
                    context.user_data['practice'] = 'Практика умножение двоичных чисел'
                    context.user_data['sort'] = 'Практика'

                    keyboard = [
                        ["Назад"],
                        ['Практика по теме']
                    ]
                    update.message.reply_text(text=theory_text.THEORY1_3,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Умножения двоичных чисел[2]':

                    context.user_data[context.user_data['sort']] = 'Умножения двоичных чисел[3]'

                    keyboard = [
                        ["Далее"]
                    ]

                    update.message.reply_text(text=theory_text.THEORY1_2_2,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-28.jpg', 'rb'))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Умножения двоичных чисел[1]':

                    context.user_data[context.user_data['sort']] = 'Умножения двоичных чисел[2]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY1_2_1,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-26.jpg', 'rb'))


                if text == "Умножения двоичных чисел":

                    context.user_data['sort'] = 'Теория'
                    context.user_data[context.user_data['sort']] = 'Умножения двоичных чисел[1]'
                    context.user_data['back'] = 'назад теория'

                    keyboard = [
                        ["Далее"],
                        ['Назад']
                    ]
                    update.message.reply_text(text=theory_text.THEORY1_1,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))


                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Возведение в квадрат[5]':

                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[6]'
                    context.user_data['back'] = 'назад теория'
                    context.user_data['practice'] = 'Практика возведение в квадрат'
                    context.user_data['sort'] = 'Практика'

                    keyboard = [
                        ["Назад"],
                        ['Практика по теме']
                    ]
                    update.message.reply_text(text=theory_text.THEORY2_6,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_6_1,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_6_2,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-33.jpg', 'rb'))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Возведение в квадрат[4]':

                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[5]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY2_5,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_5_1,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_5_2,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-31.jpg', 'rb'))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Возведение в квадрат[3]':

                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[4]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY2_4,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_4_2,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY2_4_3,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-29.jpg', 'rb'))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Возведение в квадрат[2]':

                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[3]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY2_3,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-35.jpg', 'rb'))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Возведение в квадрат[1]':

                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[2]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY2_2,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/photo_2022-01-23_21-29-37.jpg', 'rb'))

                if text == "Возведение в квадрат":

                    context.user_data['sort'] = 'Теория'
                    context.user_data[context.user_data['sort']] = 'Возведение в квадрат[1]'
                    context.user_data['back'] = 'назад теория'

                    keyboard = [
                        ["Далее"],
                        ['Назад']
                    ]

                    update.message.reply_photo(open('theory_photo_dir/4bedf30c2410ab76334d86f35aaf689c (1).png', 'rb'))

                    update.message.reply_text(text=theory_text.THEORY2_1,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Квадратные уравнения[2]':

                    context.user_data[context.user_data['sort']] = 'Квадратные уравнения[3]'
                    context.user_data['back'] = 'назад теория'
                    context.user_data['practice'] = 'Практика квадратные уравнения'
                    context.user_data['sort'] = 'Практика'

                    keyboard = [
                        ["Назад"],
                        ['Практика по теме']
                    ]
                    update.message.reply_text(text=theory_text.THEORY3_3,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_3_1,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/slide-9.jpg', 'rb'))
                    update.message.reply_text(text=theory_text.THEORY3_3_3,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_3_4,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_3_5,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_3_6,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_3_7,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))

                if text == "Далее" and context.user_data[context.user_data['sort']] == 'Квадратные уравнения[1]':

                    context.user_data[context.user_data['sort']] = 'Квадратные уравнения[2]'

                    keyboard = [
                        ["Далее"]
                    ]
                    update.message.reply_text(text=theory_text.THEORY3_2,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_2_1,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_photo(open('theory_photo_dir/diskriminant-i-korni-kvadratnogo-uravneniya.png', 'rb'))
                    update.message.reply_text(text=theory_text.THEORY3_2_3,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))
                    update.message.reply_text(text=theory_text.THEORY3_2_4,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))

                if text == "Квадратные уравнения":

                    context.user_data['sort'] = 'Теория'
                    context.user_data[context.user_data['sort']] = 'Квадратные уравнения[1]'
                    context.user_data['back'] = 'назад теория'

                    keyboard = [
                        ["Далее"],
                        ['Назад']
                    ]
                    update.message.reply_text(text=theory_text.THEORY3_1,
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))

                    update.message.reply_text(text=theory_text.THEORY3_1_1,
                                                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                                   resize_keyboard=True))



            elif update.message.text != 'Практика' and context.user_data['sort'] == 'Практика' or\
                    check_practice and context.user_data['practice'] == f'Практика{1 or 2 or 3 or 4}':

                if text == 'Назад':
                    do_start(update=update, context=context)
                    return

                if text == "Умножения двоичных чисел" or text == 'Практика по теме' and context.user_data[
                    'practice'] == 'Практика умножение двоичных чисел' or check_practice and context.user_data['practice'] == 'Практика1':
                    keyboard = [
                        ["Назад"]
                    ]
                    try:
                        if check_practice and context.user_data['practice'] == 'Практика1':
                            if text != str(context.user_data['first_number'] * context.user_data['second_number']):
                                update.message.reply_text(text=f'Неправильно, попробуйте ещё раз!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(f_ans=True, name=user_name)
                                return
                            else:
                                update.message.reply_text(text=f'Правильно, двигаемся дальше!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard,one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(t_ans=True, name=user_name)
                        else:
                            update.message.reply_text(text=f'Итак, давай поупражняемся!',
                                                      reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                       resize_keyboard=True))
                    except Exception as ex:
                        print('\033[31mLine: ', extract_tb(exc_info()[2])[0][1], '\nException: ', ex)


                    context.user_data['sort'] = 'Практика'
                    context.user_data['practice'] = 'Практика1'

                    first_number = context.user_data['first_number'] = random.randint(10, 99)
                    second_number = context.user_data['second_number'] = random.randint(first_number-3, first_number+3)
                    print(f'\033[32m{first_number*second_number}')

                    update.message.reply_text(text=f'{first_number} * {second_number}',
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))


                elif text == "Возведение в квадрат" or text == 'Практика по теме' and context.user_data[
                    'practice'] == 'Практика возведение в квадрат' or check_practice and context.user_data['practice'] == 'Практика2':

                    keyboard = [
                        ["Назад"]
                    ]
                    try:
                        if check_practice and context.user_data['practice'] == 'Практика2':
                            if text != str(context.user_data['first_number'] * context.user_data['first_number']):
                                update.message.reply_text(text=f'Неправильно, попробуйте ещё раз!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(f_ans=True, name=user_name)
                                return
                            else:
                                update.message.reply_text(text=f'Правильно, двигаемся дальше!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard,one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(t_ans=True, name=user_name)
                        else:
                            update.message.reply_text(text=f'Итак, давай поупражняемся!',
                                                      reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                       resize_keyboard=True))
                    except Exception as ex:
                        print('\033[31mLine: ', extract_tb(exc_info()[2])[0][1], '\nException: ', ex)


                    context.user_data['sort'] = 'Практика'
                    context.user_data['practice'] = 'Практика2'

                    first_number = context.user_data['first_number'] = random.randint(10, 99)
                    print(f'\033[32m{first_number*first_number}')

                    update.message.reply_text(text=f'Квадрат числа: {first_number}',
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))

                elif text == "Квадратные уравнения" or text == 'Практика по теме' and context.user_data[
                    'practice'] == 'Практика квадратные уравнения' or check_practice and context.user_data['practice'] == 'Практика3':

                    keyboard = [
                        ["Назад"]
                    ]
                    try:
                        if check_practice and context.user_data['practice'] == 'Практика3':
                            answer1 = f"{context.user_data['x1']} {context.user_data['x2']}"
                            answer2 = f"{context.user_data['x2']} {context.user_data['x1']}"
                            if text != answer1 and text != answer2:
                                update.message.reply_text(text=f'Неправильно, попробуйте ещё раз!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(f_ans=True, name=user_name)
                                return
                            else:
                                update.message.reply_text(text=f'Правильно, двигаемся дальше!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard,one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(t_ans=True, name=user_name)
                        else:
                            update.message.reply_text(text=f'Итак, давай поупражняемся! В ответ запиши корни через пробел в любом порядке',
                                                      reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                       resize_keyboard=True))

                    except Exception as ex:
                        print('\033[31mLine: ', extract_tb(exc_info()[2])[0][1], '\nException: ', ex)

                    context.user_data['sort'] = 'Практика'
                    context.user_data['practice'] = 'Практика3'

                    while True:
                        A_coefficient = random.randint(1,2)
                        B_coefficient = random.randint(-25, 25)
                        C_coefficient = random.randint(-40, 40)
                        D = B_coefficient ** 2 - 4 * A_coefficient * C_coefficient
                        sq_D = D ** 0.5
                        if D > 0 and sq_D - int(sq_D) == 0.0:
                            x1_pre = (-1 * B_coefficient + sq_D) / (2 * A_coefficient)
                            x2_pre = (-1 * B_coefficient - sq_D) / (2 * A_coefficient)
                            if abs(x1_pre) - abs(int(x1_pre)) == 0.0 and abs(x2_pre) - abs(int(x2_pre)) == 0.0:
                                x1 = context.user_data['x1'] = int(x1_pre)
                                x2 = context.user_data['x2'] = int(x2_pre)
                                print(f'\033[32m{x1, x2}')
                                break

                    mes_text = f'Найдите корни уравнения: {A_coefficient if A_coefficient == 2 else ""}x^2 ' \
                           f'{"-" if B_coefficient < 0 else "+"} {abs(B_coefficient)}x {"-" if C_coefficient < 0 else "+"}' \
                           f' {abs(C_coefficient)} = 0'
                    update.message.reply_text(text=mes_text,reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                         resize_keyboard=True))

                elif text == "Поиск процента" or text == 'Практика по теме' and context.user_data[
                    'practice'] == 'Поиск процента' or check_practice and context.user_data['practice'] == 'Практика4':

                    keyboard = [
                        ["Назад"]
                    ]
                    try:
                        if check_practice and context.user_data['practice'] == 'Практика4':
                            if text != str(context.user_data['answer_fs']):
                                update.message.reply_text(text=f'Неправильно, попробуйте ещё раз!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(f_ans=True, name=user_name)
                                return
                            else:
                                update.message.reply_text(text=f'Правильно, двигаемся дальше!',
                                    reply_markup=ReplyKeyboardMarkup(keyboard,one_time_keyboard=True,resize_keyboard=True))
                                user_name = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
                                do_statistic(t_ans=True, name=user_name)
                        else:
                            update.message.reply_text(text=f'Итак, давай поупражняемся! Не целые числа пишутся с точкой и округляются до десятых',
                                                      reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                                       resize_keyboard=True))
                    except Exception as ex:
                        print('\033[31mLine: ', extract_tb(exc_info()[2])[0][1], '\nException: ', ex)


                    context.user_data['sort'] = 'Практика'
                    context.user_data['practice'] = 'Практика4'

                    while True:
                        first_number = context.user_data['first_number'] = random.randint(10, 1000)
                        second_number = context.user_data['first_number'] = random.randint(5, 100)
                        if first_number % 5 == 0 and second_number % 5 == 0:
                            answer_fs = first_number * (second_number / 100)
                            if answer_fs - int(answer_fs) == 0.0:
                                context.user_data['answer_fs'] = round(int(answer_fs), 1)
                            else:
                                context.user_data['answer_fs'] = round(answer_fs, 1)
                            break

                    print(f'\033[32m{context.user_data["answer_fs"]}')

                    update.message.reply_text(text=f'Найдите {second_number} процентов от {first_number}',
                                              reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                               resize_keyboard=True))

    except Exception as ex:
        # print('\033[31mLine: ', extract_tb(exc_info()[2])[0][1], '\nException: ', ex)
        pass

def do_command(update, context: CallbackContext):
    if update.message.text == '/start':
        print(f'\033[32m{update.message.from_user.first_name} {update.message.from_user.last_name} подключился!')
        print(update.message.from_user.id)
        do_start(update, context)
        return

    if update.message.text == '/addven':
        context.user_data['command'] = update.message.text
        update.message.reply_text(text='Введите артикул:')

    if update.message.text == 'Сообщить о проблеме':
        context.user_data['command'] = '/addven'
        update.message.reply_text(text='Опишите вашу проблему, а так же ваши действия, которые привели вас к ошибке:')

    elif update.message.text == '/delven':
        context.user_data['command'] = update.message.text
        update.message.reply_text(text='Введите артикул для удаления:')

    elif update.message.text == '/addkey':
        context.user_data['command'] = update.message.text
        update.message.reply_text(text='Введите ключевое слово:')

    elif update.message.text == '/delkey':
        context.user_data['command'] = update.message.text
        update.message.reply_text(text='Введите ключевое слово для удаления:')

    if 'command' in context.user_data:

        if update.message.text != '/delven' and context.user_data['command'] == '/delven':
            do_words(update, context)

        elif update.message.text != '/addven' and update.message.text != 'Сообщить о проблеме' and context.user_data['command'] == '/addven':
            do_words(update, context)

        elif update.message.text != '/addkey' and context.user_data['command'] == '/addkey':
            do_words(update, context)

        elif update.message.text != '/delkey' and context.user_data['command'] == '/delkey':
            do_words(update, context)

    else:
        update.message.reply_text(text='что?...')


def do_words(update: Update, context: CallbackContext):

    if context.user_data['command'] == '/addven':
        do_add_vendor(update, context)

    elif context.user_data['command'] == '/delven':
        do_delete_vendor(update, context)

    elif context.user_data['command'] == '/addkey':
        do_add_keyword(update, context)

    elif context.user_data['command'] == '/delkey':
        do_delete_keyword(update, context)
    else:
        update.message.reply_text(text='sry...')

    context.user_data['command'] = None
    return context.user_data['command']


def do_statistic(name, t_ans = False, f_ans = False):

    if t_ans:

        for i in range(2, 100):

            if vendor.cell(column=5, row=i).value == name:
                user_val = int(vendor.cell(column=6, row=i).value) if vendor.cell(column=6,
                                                                                  row=i).value is not None else 0
                vendor.cell(column=6, row=i).value = user_val + 1
                break
        else:
            for j in range(2, 100):
                if vendor.cell(column=5, row=j).value is None:
                    vendor.cell(column=5, row=j).value = name
                    vendor.cell(column=6, row=j).value = 1
                    break
    if f_ans:

        for i in range(2, 100):

            if vendor.cell(column=5, row=i).value == name:
                user_val = int(vendor.cell(column=7, row=i).value) if vendor.cell(column=7,
                                                                                  row=i).value is not None else 0
                vendor.cell(column=7, row=i).value = user_val + 1
                break
        else:
            for j in range(2, 100):
                if vendor.cell(column=5, row=j).value is None:
                    vendor.cell(column=5, row=j).value = name
                    vendor.cell(column=7, row=j).value = 1
                    break

    return wb.save('database.xlsx')


def do_show_top_statistic(update: Update, context):
    top_list = []

    for i in range(2, 100):

        if vendor.cell(column=5, row=i).value is not None:
            user_val_t = int(vendor.cell(column=6, row=i).value) if vendor.cell(column=6, row=i).value is not None else 0
            user_list = [vendor.cell(column=5, row=i).value, user_val_t, vendor.cell(column=7, row=i).value]
            top_list.append(user_list)
            print(vendor.cell(column=5, row=i).value)
    sorted_top_list = sorted(top_list, key=itemgetter(1), reverse=True)
    print(sorted_top_list)

    update.message.reply_text(
        text='\n'.join(*sorted_top_list))

def do_show_statistic(name, update: Update, context):

    # keyboard = [
    #     ['Назад']
    # ]

    for i in range(2, 100):

        if vendor.cell(column=5, row=i).value == name:
            t_ans = vendor.cell(column=6, row=i).value if vendor.cell(column=6, row=i).value is not None else 0
            f_ans = vendor.cell(column=7, row=i).value if vendor.cell(column=7, row=i).value is not None else 0
            update.message.reply_text(
                text=f'Статистика {name}\nПравильных ответов: {t_ans}\nНеправильный ответов: {f_ans}',
                reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                 resize_keyboard=True))
            break
    else:
        update.message.reply_text(text='Прости, но я не нашел тебя в списке пользователей(',
                                  reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True,
                                                                   resize_keyboard=True))

def do_add_vendor(update: Update, context):

    for i in range(2, 100):

        if vendor.cell(column=1, row=i).value is None:
            vendor.cell(column=1, row=i).value = update.message.text
            vendor.cell(column=2, row=i).value = f'{update.message.from_user.first_name} {update.message.from_user.last_name}'
            update.message.reply_text(text='Большое спасибо, мы это исправим!')
            update.message.reply_text(reply_to_message_id='3447', text=f'{update.message.from_user.first_name} {update.message.from_user.last_name}: {update.message.text}')
            break

    return wb.save('database.xlsx')


def do_delete_vendor(update: Update, context):

    for i in range(2, 100):

        if vendor.cell(column=1, row=i).value == update.message.text:
            vendor.cell(column=1, row=i).value = None
            update.message.reply_text(text='Готово!')
            break

    return wb.save('database.xlsx')


def do_add_keyword(update: Update, context):

    for i in range(2, 100):

        if vendor.cell(column=2, row=i).value is None:
            vendor.cell(column=2, row=i).value = update.message.text
            update.message.reply_text(text='Готово!')
            break

    return wb.save('database.xlsx')


def do_delete_keyword(update: Update, context):

    for i in range(2, 100):

        if vendor.cell(column=2, row=i).value == update.message.text:
            vendor.cell(column=2, row=i).value = None
            update.message.reply_text(text='Готово!')
            break

    return wb.save('database.xlsx')

if __name__ == '__main__':
    main()
